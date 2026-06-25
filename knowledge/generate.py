#!/usr/bin/env python3
"""Generate camera feature list for any project from knowledge base.

Usage:
  python3 knowledge/generate.py --project 25131
  python3 knowledge/generate.py --project 26111 --format markdown
  python3 knowledge/generate.py --project 25111 --format excel
  python3 knowledge/generate.py --project 26111 --format all

Supports: 25111, 25131, 26111 (add more by creating devices/{project}.yaml)
"""

import argparse, json, os, sys, copy
from pathlib import Path
from collections import defaultdict

KB = Path(__file__).parent
OUT = KB / '_output'


# ===== Helpers =====

def load_json(path):
    with open(path) as f:
        return json.load(f)


def load_device_config(project):
    """Load device config from JSON."""
    path = KB / 'devices' / f'{project}.json'
    if not path.exists():
        raise FileNotFoundError(f"No device config found: {path}")
    return load_json(path)


def load_features(project):
    """Load baseline feature JSONs. For projects that inherit, load the baseline."""
    rear = load_json(KB / 'features' / 'rear-camera.json')
    front = load_json(KB / 'features' / 'front-camera.json')
    fl = load_json(KB / 'features' / 'focal-lengths.json')
    return rear, front, fl


def apply_feature_deltas(features, device_config, key='rear'):
    """Apply structured deltas from device config."""
    result = copy.deepcopy(features)
    deltas = device_config.get('feature_deltas', {})

    # Remove features
    removed = deltas.get(f'{key}_removed', [])
    if removed:
        result = [f for f in result if f['name'] not in removed and f.get('group', '') not in removed]

    # Update feature names
    updates = deltas.get(f'{key}_updates', {})
    for f in result:
        if f['name'] in updates:
            f['name'] = updates[f['name']]

    # Add new features
    additions = deltas.get(f'{key}_additions', [])
    result.extend(additions)

    return result


# ===== Markdown =====

def md_support_table(features, title, project, baseline_label):
    lines = [f'# {title}', '', f'> 项目 {project} | {baseline_label} | {len(features)} 项功能', '']

    by_mode = defaultdict(list)
    for f in features:
        by_mode[f['mode']].append(f)

    for mode in sorted(by_mode):
        items = by_mode[mode]
        cam_keys = list(items[0].get('support', {}).keys())
        cam_header = ' | '.join(cam_keys)
        header = f'| 分类 | 功能 | {cam_header} | 验证方式 |'
        sep = f'|------|------|{"|".join("------" for _ in cam_keys)}|------|'

        lines.append(f'## {mode}')
        lines.append('')
        lines.append(header)
        lines.append(sep)

        for f in items:
            def ss(v):
                if v == 'supported': return '✅'
                if v == 'unsupported': return '❌'
                return str(v)
            support_str = ' | '.join(ss(f['support'].get(c, '')) for c in cam_keys)
            verify = str(f.get('verify', ''))[:80]
            lines.append(f'| {f["category"]} | {f["group"]} > {f["name"]} | {support_str} | {verify} |')

        lines.append('')

    return '\n'.join(lines)


def md_device_specs(config, project):
    """Generate device specs table from device config."""
    lines = ['# 设备规格', '', f'> 项目 {project}', '']

    cameras = config.get('cameras', {})
    base = cameras.get('base', {})
    pro = cameras.get('pro', {})

    if base:
        lines.append(f'## {project} Base')
        lines.append('')
        cam_list = [('主摄', base.get('main', {})), ('超广角', base.get('ultrawide', {})), ('前置', base.get('front', {}))]
        _write_cam_table(lines, cam_list, base)

    if pro:
        lines.append(f'## {project} Pro')
        lines.append('')
        pro_note = pro.get('note', '')
        if pro_note:
            lines.append(f'> {pro_note}')
            lines.append('')
        cam_list = [('主摄', pro.get('main', {})), ('超广角', pro.get('ultrawide', {})),
                     ('长焦', pro.get('tele', {})), ('前置', pro.get('front', {}))]
        _write_cam_table(lines, cam_list, pro)

    # Inheritance / deltas info
    inheritance = config.get('inheritance', {})
    if inheritance:
        baseline = inheritance.get('baseline', '')
        lines.append('## 继承信息')
        lines.append('')
        lines.append(f'- 基线: {baseline}')
        for rule in inheritance.get('rules', []):
            lines.append(f'- {rule}')
        lines.append('')

    # Key deltas
    deltas_keys = [k for k in config.keys() if k.startswith('key_deltas')]
    for dk in deltas_keys:
        lines.append(f'## {dk.replace("_", " ").title()}')
        lines.append('')
        for d in config[dk]:
            lines.append(f'- {d}')
        lines.append('')

    # New features
    if config.get('new_features_p0'):
        lines.append('## P0 新增功能')
        lines.append('')
        for f in config['new_features_p0']:
            lines.append(f'- {f}')
        lines.append('')

    # Removed features
    if config.get('removed_features'):
        lines.append('## 已移除功能')
        lines.append('')
        for f in config['removed_features']:
            lines.append(f'- {f}')
        lines.append('')

    return '\n'.join(lines)


def _write_cam_table(lines, cam_list, parent):
    """Write a camera spec table row."""
    # Collect all attribute keys
    all_keys = []
    for _, cam in cam_list:
        if isinstance(cam, dict):
            all_keys.extend(cam.keys())
    all_keys = list(dict.fromkeys(all_keys))  # dedup, preserve order

    # Filter to display-worthy keys
    display_keys = [k for k in all_keys if k not in ('note', 'specs')]

    # Build header
    cam_names = [name for name, _ in cam_list]
    lines.append('| 属性 | ' + ' | '.join(cam_names) + ' |')
    lines.append('|------|' + '|'.join(['------'] * len(cam_names)) + '|')

    for key in display_keys:
        vals = []
        for _, cam in cam_list:
            if isinstance(cam, dict):
                v = cam.get(key, '—')
                if isinstance(v, bool):
                    v = '✅' if v else '❌'
                vals.append(str(v))
            else:
                vals.append(str(cam) if cam else '—')
        lines.append(f'| {key} | ' + ' | '.join(vals) + ' |')

    lines.append('')


# ===== Main =====

def main():
    parser = argparse.ArgumentParser(description='Generate camera feature list')
    parser.add_argument('--project', required=True, help='Project code (e.g., 25131, 26111)')
    parser.add_argument('--format', default='all', choices=['markdown', 'excel', 'all'],
                        help='Output format (default: all)')
    args = parser.parse_args()

    project = args.project
    fmt = args.format

    OUT.mkdir(parents=True, exist_ok=True)

    # Load device config
    config = load_device_config(project)

    # Load baseline features
    rear, front, fl = load_features(project)

    # Get project info (handles both formats: dict in YAML, string in JSON)
    proj_info = config.get('project', {})
    if isinstance(proj_info, str):
        code = proj_info
    else:
        code = proj_info.get('code', project)

    # Determine baseline label
    inheritance = config.get('inheritance', {})
    baseline = inheritance.get('baseline', '')
    baseline_label = f'{baseline} → {project} 继承基线' if baseline else '基准项目'

    # Apply deltas from device config
    rear_features = rear['features']
    front_features = front['features']
    if config.get('feature_deltas'):
        rear_features = apply_feature_deltas(rear_features, config, 'rear')
        front_features = apply_feature_deltas(front_features, config, 'front')

    # Generate markdown
    if fmt in ('markdown', 'all'):
        md = []
        md.append(f'# {code} 相机功能列表\n')
        md.append(f'> 基线: {baseline_label}')
        md.append(f'> 自动生成 | 后置 {len(rear_features)} 项 + 前置 {len(front_features)} 项')
        if not config.get('feature_deltas'):
            md.append(f'> ⚠️ 未定义 feature_deltas，输出为基线 feature 列表')
        md.append('')
        md.append('---\n')
        md.append(md_support_table(rear_features, '后置摄像头功能列表', code, baseline_label))
        md.append('\n---\n')
        md.append(md_support_table(front_features, '前置摄像头功能列表', code, baseline_label))
        md.append('\n---\n')
        md.append(md_device_specs(config, code))

        out = '\n'.join(md)
        path = OUT / f'features-{project}.md'
        path.write_text(out)
        print(f'Markdown: {path} ({len(out):,} chars)')
        print(f'  Rear: {len(rear_features)} features')
        print(f'  Front: {len(front_features)} features')

    # Generate excel (reuse existing logic for backward compat)
    if fmt in ('excel', 'all'):
        generate_excel(project, rear_features, front_features, config)


def generate_excel(project, rear_features, front_features, config):
    """Generate Excel output."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        print("Need openpyxl: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1C1A16', end_color='1C1A16', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Rear Camera
    ws = wb.active
    ws.title = '后置摄像头'

    # Detect camera keys from features
    cam_keys = ['ultrawide', 'main', 'macro']
    if rear_features:
        first = rear_features[0]
        cam_keys = list(first.get('support', {}).keys()) or cam_keys

    headers = ['模式', '分类', '功能组', '子功能'] + cam_keys + ['验证方式']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for ri, f in enumerate(rear_features, 2):
        ws.cell(row=ri, column=1, value=f['mode']).border = thin_border
        ws.cell(row=ri, column=2, value=f['category']).border = thin_border
        ws.cell(row=ri, column=3, value=f['group']).border = thin_border
        ws.cell(row=ri, column=4, value=f['name']).border = thin_border
        for ci, ck in enumerate(cam_keys, 5):
            v = f['support'].get(ck, '')
            ws.cell(row=ri, column=ci, value='√' if v == 'supported' else ('×' if v == 'unsupported' else str(v))).border = thin_border
        ws.cell(row=ri, column=len(headers), value=f.get('verify', '')).border = thin_border

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    # Front Camera
    ws2 = wb.create_sheet('前置摄像头')
    headers2 = ['模式', '分类', '功能组', '子功能', '前置', '验证方式']
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for ri, f in enumerate(front_features, 2):
        ws2.cell(row=ri, column=1, value=f['mode']).border = thin_border
        ws2.cell(row=ri, column=2, value=f['category']).border = thin_border
        ws2.cell(row=ri, column=3, value=f['group']).border = thin_border
        ws2.cell(row=ri, column=4, value=f['name']).border = thin_border
        v = f['support'].get('front', '')
        ws2.cell(row=ri, column=5, value='√' if v == 'supported' else ('×' if v == 'unsupported' else str(v))).border = thin_border
        ws2.cell(row=ri, column=6, value=f.get('verify', '')).border = thin_border

    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = 'A2'

    out_path = OUT / f'features-{project}.xlsx'
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f'Excel: {out_path}')


if __name__ == '__main__':
    main()
